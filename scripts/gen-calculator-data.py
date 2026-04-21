"""
통신사 요금 계산기 Excel — 3사 분리 계산기 (각 통신사 전용 시트)
출처: docs/calculator.html D 객체 (2026-04-21)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os, shutil

OUT = '/Users/vinsenzo/bongi-mobile/docs/bongi-calculator-data.xlsx'
SCRIPT_OUT = '/Users/vinsenzo/bongi-mobile/scripts/gen-calculator-data.py'

# ─── 데이터 ───
SKT_TVS = [['TV 없음',0,0,0],['B tv 이코노미',182,12100,2200],['B tv 스탠다드',236,15400,2200],
           ['B tv 올',252,18700,2200],['B tv 스탠다드 플러스',222,23100,2200],['B tv 올 플러스',252,24200,2200],
           ['B tv 스탠다드 넷플릭스',222,27700,2200],['B tv 올 넷플릭스',222,30200,2200],
           ['B tv 스탠다드 넷플릭스 프리미엄',252,30700,2200],['B tv 올 넷플릭스 프리미엄',252,33200,2200]]
KT_TVS = [['TV 없음',0,0,0],['지니TV 베이직',238,14740,2640],['지니TV 라이트',240,15840,2640],
          ['지니TV 에센스',263,20240,3740],['지니TV 모든G',250,21340,4400],['지니TV 디즈니+모든G',250,28100,6600]]
LGU_TVS = [['TV 없음',0,0,0],['U+tv 실속형',217,15400,2200],['U+tv 기본형',223,16500,2200],
           ['U+tv 프리미엄',252,18700,2200],['U+tv 프리미엄 VOD',257,24200,5500]]

HEAD_FONT = Font(name='Noto Sans KR', size=11, bold=True, color='ffffff')
BODY_FONT = Font(name='Noto Sans KR', size=10)
MONO = Font(name='SF Mono', size=10)
BORDER = Border(left=Side(style='thin', color='cccccc'), right=Side(style='thin', color='cccccc'),
                top=Side(style='thin', color='cccccc'), bottom=Side(style='thin', color='cccccc'))
HEAD_FILL = PatternFill('solid', fgColor='1a2744')
SKT_FILL = PatternFill('solid', fgColor='fef2f2')
KT_FILL = PatternFill('solid', fgColor='eff6ff')
LGU_FILL = PatternFill('solid', fgColor='fdf2f8')
INPUT_FILL = PatternFill('solid', fgColor='fef3c7')
RESULT_FILL = PatternFill('solid', fgColor='fee2e2')

def head(c):
    c.fill = HEAD_FILL; c.font = HEAD_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER

def body(c, fill=None):
    c.font = BODY_FONT; c.border = BORDER
    c.alignment = Alignment(vertical='center')
    if fill: c.fill = fill

def table(ws, r, headers, rows, fill=None):
    for i, h in enumerate(headers):
        head(ws.cell(row=r, column=i+1, value=h))
    for ri, row in enumerate(rows):
        for ci, v in enumerate(row):
            body(ws.cell(row=r+1+ri, column=ci+1, value=v), fill)
    return r + 1 + len(rows)

def title(ws, r, t, color='1a2744'):
    c = ws.cell(row=r, column=1, value=t)
    c.font = Font(name='Noto Sans KR', size=14, bold=True, color=color)
    ws.row_dimensions[r].height = 24
    return r + 2

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════
# INDEX
# ═══════════════════════════════════════════════
idx = wb.active
idx.title = 'INDEX'
idx.column_dimensions['A'].width = 30
idx.column_dimensions['B'].width = 55
idx.column_dimensions['C'].width = 15
idx['A1'] = '🧮 봉이모바일 — 3사 통신 요금 계산기'
idx['A1'].font = Font(name='Noto Sans KR', size=18, bold=True, color='1a2744')
idx.merge_cells('A1:C1')
idx['A2'] = '출처: docs/calculator.html D 객체 · 2026-04-21'
idx['A2'].font = Font(name='Noto Sans KR', size=11, color='666666', italic=True)
idx.merge_cells('A2:C2')

r = title(idx, 4, '📋 시트 구조')
sheets_info = [
    ['🧮 SKT 계산기', 'SKT 전용 요즘가족결합 계산기 (드롭다운)', '★ 계산'],
    ['🧮 KT 계산기', 'KT 전용 총액/정액/프리미엄 계산기', '★ 계산'],
    ['🧮 LG U+ 계산기', 'LG U+ 전용 참쉬운/투게더 계산기', '★ 계산'],
    ['01_Internet', '인터넷 단독 요금', '데이터'],
    ['02_TV', 'TV 상품 (3사 전체)', '데이터'],
    ['03_TV_Internet', 'TV 결합 시 인터넷 (WiFi 유/무)', '데이터'],
    ['04_SetTop', '셋톱박스 (기본값 설명 포함)', '데이터'],
    ['05_WiFi', 'WiFi (기본값 설명 포함)', '데이터'],
    ['06_Install', '설치비', '데이터'],
    ['07_Gift', '사은품', '데이터'],
    ['08_Cards', '제휴카드 (26종)', '데이터'],
    ['10_SKT_Bundle', 'SKT 요즘가족결합 (15조합)', '결합'],
    ['11_KT_Total', 'KT 총액결합 (6구간)', '결합'],
    ['12_KT_Fixed', 'KT 정액결합 (4구간)', '결합'],
    ['13_KT_Premium', 'KT 💎 프리미엄 요금제 카탈로그', '결합'],
    ['14_LGU_Chweyswun', 'LG U+ 참쉬운가족결합', '결합'],
    ['15_LGU_Together', 'LG U+ 투게더 결합', '결합'],
    ['20_Schema', 'TypeScript 인터페이스', '개발'],
    ['30_Formula', '계산 공식 (REPLACE vs STACK)', '개발'],
]
r = table(idx, r, ['시트명', '내용', '분류'], sheets_info)

r += 2
c = idx.cell(row=r, column=1, value='⚡ 사용법')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='ef4444')
idx.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
r += 1
for line in [
    '1. 🧮 SKT / KT / LG U+ 계산기 중 고객 통신사 시트 선택',
    '2. 노란 셀 클릭 → 드롭다운에서 속도/TV/WiFi/결합 선택',
    '3. 하단에 단계별 계산 결과 + 🎉 혜택가 자동 표시',
    '4. 각 행에 출처 데이터 시트 주석 (→ 01_Internet 등)',
]:
    c = idx.cell(row=r, column=1, value=line)
    c.font = BODY_FONT
    idx.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

r += 2
c = idx.cell(row=r, column=1, value='🔗 참고자료 (원본 출처)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='2563eb')
idx.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
r += 1
refs = [
    ('SK 상품 전체 안내 (요즘가족결합)', 'https://www.100mb.kr/01_product/sk.php'),
    ('KT 프리미엄 가족결합 상세 해설', 'https://www.100mb.kr/bbs/board.php?bo_table=information&wr_id=11986'),
    ('LG U+ 상품 전체 안내 (참쉬운/투게더)', 'https://www.100mb.kr/01_product/lg.php'),
    ('봉이모바일 통합 계산기 (원본 데이터 소스)', 'https://bongi-mobile-production.up.railway.app/docs/calculator.html'),
    ('GitHub 리포지토리', 'https://github.com/vinsenzo83/bongi-mobile'),
]
for label, url in refs:
    c = idx.cell(row=r, column=1, value=f'🔗 {label}')
    c.hyperlink = url
    c.font = Font(name='Noto Sans KR', size=10, color='2563eb', underline='single')
    c2 = idx.cell(row=r, column=2, value=url)
    c2.font = Font(name='SF Mono', size=9, color='666666')
    idx.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    r += 1

# ═══════════════════════════════════════════════
# 계산기 공통 헬퍼
# ═══════════════════════════════════════════════
def frow(ws, r, label, formula, fill=None, bold=False, color=None, comment='', fmt='#,##0"원"'):
    ws.cell(row=r, column=1, value=label).font = Font(name='Noto Sans KR', size=10, bold=bold)
    c = ws.cell(row=r, column=2, value=formula)
    c.font = Font(name='SF Mono', size=11, bold=bold, color=color or '1a2744')
    c.border = BORDER
    c.alignment = Alignment(horizontal='right')
    c.number_format = fmt
    if fill: c.fill = fill
    if comment:
        cc = ws.cell(row=r, column=3, value=comment)
        cc.font = Font(name='Noto Sans KR', size=9, color='666666', italic=True)
        cc.alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    return r + 1

def input_row(ws, r, label, default, hint='', opts=None):
    ws.cell(row=r, column=1, value=label).font = Font(name='Noto Sans KR', size=11, bold=True)
    c = ws.cell(row=r, column=2, value=default)
    c.font = Font(name='SF Mono', size=12, bold=True, color='92400e')
    c.fill = INPUT_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal='center', vertical='center')
    if opts:
        dv = DataValidation(type='list', formula1=f'"{",".join(opts)}"', allow_blank=False)
        ws.add_data_validation(dv); dv.add(c)
    if hint:
        h = ws.cell(row=r, column=3, value=hint)
        h.font = Font(name='Noto Sans KR', size=9, italic=True, color='666666')
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    return r + 1

def calc_header(ws, title, color):
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name='Noto Sans KR', size=18, bold=True, color='ffffff')
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 36

# ═══════════════════════════════════════════════
# 🧮 SKT 계산기
# ═══════════════════════════════════════════════
ws = wb.create_sheet('🧮 SKT 계산기')
for col, w in zip('ABCDE', [28, 22, 40, 20, 15]):
    ws.column_dimensions[col].width = w
calc_header(ws, '🧮 SKT 요즘가족결합 계산기', 'ef4444')

r = 3
c = ws.cell(row=r, column=1, value='📝 입력 (노란 셀 클릭 → 드롭다운)')
c.font = Font(name='Noto Sans KR', size=12, bold=True, color='1a2744')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

skt_tv_names = [t[0] for t in SKT_TVS]
r = input_row(ws, r, '속도', '500M', '100M / 500M / 1G', ['100M','500M','1G'])
r = input_row(ws, r, 'TV 상품', 'B tv 올', 'SKT B tv 9종', skt_tv_names)
r = input_row(ws, r, 'WiFi 사용', 'N', 'Y=+1,100원 / N=미사용', ['Y','N'])
r = input_row(ws, r, '휴대폰 회선수', 3, '0=결합없음 / 1~5', ['0','1','2','3','4','5'])
r += 1

# 자동 계산
c = ws.cell(row=r, column=1, value='🧮 자동 계산')
c.font = Font(name='Noto Sans KR', size=12, bold=True, color='1a2744')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

SP, TVN, WIFI, LINES = 'B4', 'B5', 'B6', 'B7'

# 인덱스 변환
r = frow(ws, r, 'TV 인덱스 (내부 변환)', f'=MATCH({TVN},{{"TV 없음","B tv 이코노미","B tv 스탠다드","B tv 올","B tv 스탠다드 플러스","B tv 올 플러스","B tv 스탠다드 넷플릭스","B tv 올 넷플릭스","B tv 스탠다드 넷플릭스 프리미엄","B tv 올 넷플릭스 프리미엄"}},0)',
    comment='드롭다운 이름 → 배열 번호 (1=TV 없음, 2~10=상품) · 수식 조회용', fmt='General')
TV_IDX = r - 1

r = frow(ws, r, 'TV 선택 여부 (내부 변환)', f'=IF({TVN}="TV 없음",0,1)',
    comment='0=TV 없음 (인터넷 단독) / 1=TV 결합 · 수식 분기용', fmt='General')
HAS_TV = r - 1

r += 1
c = ws.cell(row=r, column=1, value='📊 데이터 조회')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
r += 1

r = frow(ws, r, '인터넷 단독가',
    f'=IF({SP}="100M",22000,IF({SP}="500M",33000,38500))',
    comment='→ 01_Internet · SKT 단독가')
SOLO = r - 1

r = frow(ws, r, 'TV 상품 가격 (p)',
    f'=IF({HAS_TV}=0,0,CHOOSE(B{TV_IDX}-1,12100,15400,18700,23100,24200,27700,30200,30700,33200))',
    comment='→ 02_TV · SKT TV 9종')
ws.cell(row=r-1, column=2).value = f'=IF(B{HAS_TV}=0,0,CHOOSE(B{TV_IDX}-1,12100,15400,18700,23100,24200,27700,30200,30700,33200))'
TV_P = r - 1

r = frow(ws, r, 'TV 결합할인 (dc)',
    f'=IF(B{HAS_TV}=0,0,2200)',
    comment='→ 02_TV · SKT 전 상품 -2,200 균일')
TV_DC = r - 1

r = frow(ws, r, 'WiFi 추가금',
    f'=IF({WIFI}="Y",1100,0)',
    comment='→ 05_WiFi · SKT GIGA WiFi (기본) 속도 균일 1,100원')
WIFI_FEE = r - 1

r = frow(ws, r, 'tvInternetNoWifi',
    f'=IF({SP}="100M",19800,IF({SP}="500M",27500,33000))',
    comment='→ 03_TV_Internet · WiFi 미포함 결합가')
TV_INET_N = r - 1

r = frow(ws, r, 'tvInternetWithWifi',
    f'=IF({SP}="100M",22000,IF({SP}="500M",28600,34100))',
    comment='→ 03_TV_Internet · WiFi 포함 결합가')
TV_INET_W = r - 1

r = frow(ws, r, '셋톱박스 (스마트3)', '=4400',
    comment='→ 04_SetTop · 기본 모델 (매장 판매 1위)')
SETTOP = r - 1

r += 1
# 경로 A
c = ws.cell(row=r, column=1, value='【경로 A】 결합없음 (휴대폰 결합 X)')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

r = frow(ws, r, '  월 기본요금 (요즘우리집 · 휴대폰결합 없을 때)',
    f'=IF(B{HAS_TV}=0,B{SOLO}+B{WIFI_FEE},IF({WIFI}="Y",B{TV_INET_W},B{TV_INET_N})+B{TV_P}-B{TV_DC}+B{SETTOP})',
    SKT_FILL, True,
    comment='TV 결합: tvInternet(WiFi반영) + (p-dc) + 셋톱 / TV 없음: 단독+WiFi')
PATH_A = r - 1

r += 1
# 경로 B
c = ws.cell(row=r, column=1, value='【경로 B】 요즘가족결합 (휴대폰 결합 · 단독가 REPLACE)')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='ef4444')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

r = frow(ws, r, '  요즘가족 인터넷 할인',
    f'=IF({LINES}=0,0,IF({SP}="100M",4400,IF({SP}="500M",11000,13200)))',
    comment='→ 10_SKT_Bundle · 단독가 기준 차감 (요즘우리집 REPLACE)')
FAM_INET = r - 1

r = frow(ws, r, '  IPTV 추가 (TV+결합 시)',
    f'=IF(AND({LINES}>0,B{HAS_TV}=1),1100,0)',
    comment='TV 결합 시에만 추가 -1,100원')
FAM_IPTV = r - 1

r = frow(ws, r, '  휴대폰 인당 할인',
    f'=IF({LINES}=0,0,IF({SP}="100M",CHOOSE({LINES},3500,3500,6000,4500,3600),CHOOSE({LINES},3500,3500,6000,6000,4800)))',
    comment='→ 10_SKT_Bundle · 속도+회선수별 인당')
MOB_PER = r - 1

r = frow(ws, r, '  휴대폰 총 할인',
    f'=B{MOB_PER}*{LINES}',
    comment='인당 × 회선수 (휴대폰 고지서 별도 차감)')
MOB_TOTAL = r - 1

r += 1
r = frow(ws, r, '  → 인터넷 (단독 - 요즘가족할인)',
    f'=IF({LINES}=0,"-",B{SOLO}+B{WIFI_FEE}-B{FAM_INET})',
    SKT_FILL)
r = frow(ws, r, '  → TV (p - dc - iptv)',
    f'=IF({LINES}=0,"-",IF(B{HAS_TV}=0,0,B{TV_P}-B{TV_DC}-B{FAM_IPTV}))',
    SKT_FILL)
r = frow(ws, r, '  → 셋톱 (TV 있을 때만)',
    f'=IF({LINES}=0,"-",IF(B{HAS_TV}=0,0,B{SETTOP}))',
    SKT_FILL)

r = frow(ws, r, '  = 인터넷+TV 월 실납부 (결합 적용)',
    f'=IF({LINES}=0,"-",(B{SOLO}+B{WIFI_FEE}-B{FAM_INET})+IF(B{HAS_TV}=0,0,B{TV_P}-B{TV_DC}-B{FAM_IPTV}+B{SETTOP}))',
    SKT_FILL, True)
BUNDLE_FEE = r - 1

r = frow(ws, r, '  - 휴대폰 고지서 별도 차감',
    f'=IF({LINES}=0,"-",-B{MOB_TOTAL})', SKT_FILL)

r = frow(ws, r, '🎉 혜택가 (최종 월요금)',
    f'=IF({LINES}=0,B{PATH_A},B{BUNDLE_FEE}-B{MOB_TOTAL})',
    RESULT_FILL, True, 'ef4444',
    comment='결합없음이면 경로 A 값 / 결합 있으면 경로 B - 휴대폰할인')
c = ws.cell(row=r-1, column=2)
c.font = Font(name='SF Mono', size=16, bold=True, color='ef4444')

r += 1
c = ws.cell(row=r, column=1, value='🎁 부가 혜택')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
r += 1
r = frow(ws, r, '  설치비 (1회성)',
    f'=IF(B{HAS_TV}=0,36300,56100)',
    comment='→ 06_Install')
r = frow(ws, r, '  사은품 (속도×TV결합)',
    f'=IF(B{HAS_TV}=0,IF({SP}="100M",110000,170000),CHOOSE(MATCH({SP},{{"100M","500M","1G"}},0),400000,430000,490000))',
    comment='→ 07_Gift')

# ═══════════════════════════════════════════════
# 🧮 KT 계산기
# ═══════════════════════════════════════════════
ws = wb.create_sheet('🧮 KT 계산기')
for col, w in zip('ABCDE', [28, 22, 40, 20, 15]):
    ws.column_dimensions[col].width = w
calc_header(ws, '🧮 KT 총액/정액 결합 계산기', '2563eb')

r = 3
c = ws.cell(row=r, column=1, value='📝 입력')
c.font = Font(name='Noto Sans KR', size=12, bold=True, color='1a2744')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

kt_tv_names = [t[0] for t in KT_TVS]
r = input_row(ws, r, '속도', '500M', '100M / 500M / 1G', ['100M','500M','1G'])
r = input_row(ws, r, 'TV 상품', '지니TV 에센스', 'KT 지니TV 5종', kt_tv_names)
r = input_row(ws, r, 'WiFi 사용', 'Y', 'Y=1,100원 (1G는 무료) / N', ['Y','N'])
r = input_row(ws, r, '결합 종류', '총액결합', '결합없음/총액/정액/💎프리미엄', ['결합없음','총액결합','정액결합','💎 프리미엄 가족결합'])
r = input_row(ws, r, '총액결합 구간 (1~6)', 4, '1=22K↓, 2=22K↑, 3=64.9K↑, 4=108.9K↑, 5=141.9K↑, 6=174.9K↑', ['1','2','3','4','5','6'])
r = input_row(ws, r, '정액결합 구간 (1~4)', 2, '1=37K↓, 2=37K↑, 3=61K↑, 4=77K↑', ['1','2','3','4'])
r = input_row(ws, r, '💎 프리미엄 요금제', '89,000원', '77K+ 요금제 (드롭다운)',
    ['77,000원','80,000원','87,890원','89,000원','90,000원','100,000원','109,000원','110,000원','130,000원'])
r = input_row(ws, r, '💎 프리미엄 회선수', 2, '대표자 제외 구성원 77K+ 회선수 (최소 1명)', ['1','2','3','4'])
r += 1

c = ws.cell(row=r, column=1, value='🧮 자동 계산')
c.font = Font(name='Noto Sans KR', size=12, bold=True, color='1a2744')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

SP, TVN, WIFI, BUNDLE, R_T, R_F, PREM_PLAN, PREM_LINES = 'B4', 'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11'

r = frow(ws, r, 'TV 인덱스 (내부 변환)',
    f'=MATCH({TVN},{{"TV 없음","지니TV 베이직","지니TV 라이트","지니TV 에센스","지니TV 모든G","지니TV 디즈니+모든G"}},0)',
    comment='드롭다운 이름 → 배열 번호 (1=없음, 2~6=상품) · 수식 조회용', fmt='General')
TV_IDX = r - 1

r = frow(ws, r, 'TV 선택 여부 (내부 변환)', f'=IF({TVN}="TV 없음",0,1)',
    comment='0=TV 없음 / 1=TV 결합 · 수식 분기용', fmt='General')
HAS_TV = r - 1

r += 1
c = ws.cell(row=r, column=1, value='📊 데이터 조회')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
r += 1

r = frow(ws, r, '인터넷 단독가',
    f'=IF({SP}="100M",22000,IF({SP}="500M",33000,38500))',
    comment='→ 01_Internet')
SOLO = r - 1

r = frow(ws, r, 'TV 가격 (p)',
    f'=IF(B{HAS_TV}=0,0,CHOOSE(B{TV_IDX}-1,14740,15840,20240,21340,28100))',
    comment='→ 02_TV · KT 지니TV')
TV_P = r - 1

r = frow(ws, r, 'TV 결합할인 (dc)',
    f'=IF(B{HAS_TV}=0,0,CHOOSE(B{TV_IDX}-1,2640,2640,3740,4400,6600))',
    comment='→ 02_TV · KT 상품별 차등')
TV_DC = r - 1

r = frow(ws, r, 'WiFi 추가금 (속도별)',
    f'=IF({WIFI}="Y",IF({SP}="1G",0,1100),0)',
    comment='→ 05_WiFi · KT WAVE2 (기본) · 1G는 무료')
WIFI_FEE = r - 1

r = frow(ws, r, 'tvInternetNoWifi',
    f'=IF({SP}="100M",22000,IF({SP}="500M",27500,33000))',
    comment='→ 03_TV_Internet')
TV_INET_N = r - 1

r = frow(ws, r, 'tvInternetWithWifi',
    f'=IF({SP}="100M",23100,IF({SP}="500M",28600,33000))',
    comment='→ 03_TV_Internet · 1G는 동일')
TV_INET_W = r - 1

r = frow(ws, r, '셋톱박스 (기가지니3)', '=4400',
    comment='→ 04_SetTop · 기본 모델 · 고객 선호 1위')
SETTOP = r - 1

r += 1
c = ws.cell(row=r, column=1, value='【경로 A】 결합없음 (기본 TV 결합)')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

r = frow(ws, r, '  월 기본요금 (기본 TV결합 · 휴대폰결합 없을 때)',
    f'=IF(B{HAS_TV}=0,B{SOLO}+B{WIFI_FEE},IF({WIFI}="Y",B{TV_INET_W},B{TV_INET_N})+B{TV_P}-B{TV_DC}+B{SETTOP})',
    KT_FILL, True,
    comment='KT 기본 TV결합 할인 자동 적용 (tvInternetNoWifi/WithWifi)')
PATH_A = r - 1

r += 1
c = ws.cell(row=r, column=1, value='【경로 B】 결합 적용 (STACK — 기본 TV결합 + 추가할인)')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='2563eb')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

r = frow(ws, r, '  [총액] 인터넷 추가 할인',
    f'=IF({BUNDLE}="총액결합",IF({SP}="100M",CHOOSE({R_T},1650,3300,5500,5500,5500,5500),CHOOSE({R_T},2200,5500,5500,5500,5500,5500)),0)',
    comment='→ 11_KT_Total · STACK 중복 적용')
TOTAL_INET = r - 1

r = frow(ws, r, '  [총액] 휴대폰 할인',
    f'=IF({BUNDLE}="총액결합",IF({SP}="100M",CHOOSE({R_T},0,0,3300,14300,18700,23100),CHOOSE({R_T},0,0,5500,16610,22110,27610)),0)',
    comment='→ 11_KT_Total · 휴대폰 고지서 별도')
TOTAL_MOB = r - 1

r = frow(ws, r, '  [정액] 인터넷 할인',
    f'=IF({BUNDLE}="정액결합",5500,0)',
    comment='→ 12_KT_Fixed · 전구간 5,500 고정')
FIXED_INET = r - 1

r = frow(ws, r, '  [정액] 휴대폰 할인',
    f'=IF({BUNDLE}="정액결합",CHOOSE({R_F},0,3000,5000,7000),0)',
    comment='→ 12_KT_Fixed · 4구간')
FIXED_MOB = r - 1

# 💎 프리미엄 가족결합
r = frow(ws, r, '  [💎프리미엄] 인터넷 할인 (고정)',
    f'=IF({BUNDLE}="💎 프리미엄 가족결합",5500,0)',
    comment='→ 13_KT_Premium · 인터넷 또는 대표자 휴대폰 고정 -5,500')
PREM_INET = r - 1

r = frow(ws, r, '  [💎프리미엄] 대표자 총액결합 (합산=요금제1명)',
    f'=IF({BUNDLE}="💎 프리미엄 가족결합",IF({SP}="100M",CHOOSE({R_T},0,0,3300,14300,18700,23100),CHOOSE({R_T},0,0,5500,16610,22110,27610)),0)',
    comment='총액결합 구간 활용 (대표자 1명 기준)')
PREM_REP_MOB = r - 1

r = frow(ws, r, '  [💎프리미엄] 요금제별 할인액',
    f'=IF({BUNDLE}="💎 프리미엄 가족결합",CHOOSE(MATCH({PREM_PLAN},{{"77,000원","80,000원","87,890원","89,000원","90,000원","100,000원","109,000원","110,000원","130,000원"}},0),19250,20000,22000,22250,22500,25000,27500,27500,32500),0)',
    comment='→ 13_KT_Premium · planCatalog dc')
PREM_DC = r - 1

r = frow(ws, r, '  [💎프리미엄] 구성원 총 할인 (dc × 회선수)',
    f'=B{PREM_DC}*IF({BUNDLE}="💎 프리미엄 가족결합",{PREM_LINES},0)',
    comment='77K+ 구성원만 해당 (대표자 제외)')
PREM_MOB_TOTAL = r - 1

r += 1
# 인터넷+TV 월 실납부 — 결합 종류별 분기
r = frow(ws, r, '  = 인터넷+TV 월 실납부',
    f'=IF({BUNDLE}="💎 프리미엄 가족결합",B{PATH_A}-B{PREM_INET},B{PATH_A}-B{TOTAL_INET}-B{FIXED_INET})',
    KT_FILL, True,
    comment='프리미엄: -5,500 고정 / 총액+정액: 해당 할인')
BUNDLE_FEE = r - 1

r = frow(ws, r, '  - 휴대폰 고지서 별도 차감',
    f'=-(B{TOTAL_MOB}+B{FIXED_MOB}+B{PREM_REP_MOB}+B{PREM_MOB_TOTAL})', KT_FILL,
    comment='총액/정액 + 프리미엄(대표자 총액 + 구성원 프리미엄)')

r = frow(ws, r, '🎉 혜택가 (최종 월요금)',
    f'=IF({BUNDLE}="결합없음",B{PATH_A},B{BUNDLE_FEE}-(B{TOTAL_MOB}+B{FIXED_MOB}+B{PREM_REP_MOB}+B{PREM_MOB_TOTAL}))',
    RESULT_FILL, True, '2563eb')
c = ws.cell(row=r-1, column=2)
c.font = Font(name='SF Mono', size=16, bold=True, color='2563eb')

r += 1
c = ws.cell(row=r, column=1, value='🎁 부가')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
r += 1
r = frow(ws, r, '  설치비', f'=IF(B{HAS_TV}=0,36000,56200)', comment='→ 06_Install')
r = frow(ws, r, '  사은품', f'=IF(B{HAS_TV}=0,IF({SP}="100M",90000,140000),CHOOSE(MATCH({SP},{{"100M","500M","1G"}},0),370000,450000,450000))', comment='→ 07_Gift')

# ═══════════════════════════════════════════════
# 🧮 LG U+ 계산기
# ═══════════════════════════════════════════════
ws = wb.create_sheet('🧮 LG U+ 계산기')
for col, w in zip('ABCDE', [28, 22, 40, 20, 15]):
    ws.column_dimensions[col].width = w
calc_header(ws, '🧮 LG U+ 참쉬운/투게더 결합 계산기', 'e40981')

r = 3
c = ws.cell(row=r, column=1, value='📝 입력')
c.font = Font(name='Noto Sans KR', size=12, bold=True, color='1a2744')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

lgu_tv_names = [t[0] for t in LGU_TVS]
r = input_row(ws, r, '속도', '500M', '100M / 500M / 1G', ['100M','500M','1G'])
r = input_row(ws, r, 'TV 상품', 'U+tv 기본형', 'LG U+tv 4종', lgu_tv_names)
r = input_row(ws, r, '결합 종류', '참쉬운 가족결합', '결합없음/참쉬운/투게더', ['결합없음','참쉬운 가족결합','투게더 결합'])
r = input_row(ws, r, '요금구간 (참쉬운, 1~3)', 2, '1=69K미만, 2=69K+, 3=88K+', ['1','2','3'])
r = input_row(ws, r, '결합 회선수', 2, '참쉬운 2~4+ / 투게더 2~5', ['2','3','4','5'])
r += 1

c = ws.cell(row=r, column=1, value='🧮 자동 계산')
c.font = Font(name='Noto Sans KR', size=12, bold=True, color='1a2744')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

SP, TVN, BUNDLE, RANGE, LINES = 'B4', 'B5', 'B6', 'B7', 'B8'

r = frow(ws, r, 'TV 인덱스 (내부 변환)',
    f'=MATCH({TVN},{{"TV 없음","U+tv 실속형","U+tv 기본형","U+tv 프리미엄","U+tv 프리미엄 VOD"}},0)',
    comment='드롭다운 이름 → 배열 번호 (1=없음, 2~5=상품) · 수식 조회용', fmt='General')
TV_IDX = r - 1

r = frow(ws, r, 'TV 선택 여부 (내부 변환)', f'=IF({TVN}="TV 없음",0,1)',
    comment='0=TV 없음 / 1=TV 결합 · 수식 분기용', fmt='General')
HAS_TV = r - 1

r += 1
c = ws.cell(row=r, column=1, value='📊 데이터 조회')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
r += 1

r = frow(ws, r, '인터넷 단독가',
    f'=IF({SP}="100M",22000,IF({SP}="500M",33000,38500))',
    comment='→ 01_Internet')
SOLO = r - 1

r = frow(ws, r, 'TV 가격 (p)',
    f'=IF(B{HAS_TV}=0,0,CHOOSE(B{TV_IDX}-1,15400,16500,18700,24200))',
    comment='→ 02_TV · LG U+tv')
TV_P = r - 1

r = frow(ws, r, 'TV 결합할인 (dc)',
    f'=IF(B{HAS_TV}=0,0,CHOOSE(B{TV_IDX}-1,2200,2200,2200,5500))',
    comment='→ 02_TV · 전 상품 -2,200 (VOD만 -5,500)')
TV_DC = r - 1

r = frow(ws, r, 'WiFi 추가금', '=0',
    comment='→ 05_WiFi · LG U+ 전속도 무료 (기가와이파이 기본)')

r = frow(ws, r, '셋톱박스 (U+tv UHD4)', '=4400',
    comment='→ 04_SetTop · 기본 모델 · AI음향 기술')
SETTOP = r - 1

r += 1
c = ws.cell(row=r, column=1, value='【경로 A】 결합없음')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

r = frow(ws, r, '  월 기본요금',
    f'=IF(B{HAS_TV}=0,B{SOLO},IF({SP}="100M",22000,IF({SP}="500M",27500,33000))+B{TV_P}-B{TV_DC}+B{SETTOP})',
    LGU_FILL, True,
    comment='TV 없으면 단독가 (WiFi 무료) / TV 있으면 tvInternet + (p-dc) + 셋톱')
PATH_A = r - 1

r += 1
c = ws.cell(row=r, column=1, value='【경로 B】 결합 적용 (REPLACE — 단독가 기준)')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='e40981')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

r = frow(ws, r, '  [참쉬운] 인터넷 할인',
    f'=IF({BUNDLE}="참쉬운 가족결합",IF({SP}="100M",5500,IF({SP}="500M",9900,13200)),0)',
    comment='→ 14_LGU_Chweyswun · 단독가 REPLACE')
CHW_INET = r - 1

r = frow(ws, r, '  [참쉬운] 휴대폰 인당 할인',
    f'=IF({BUNDLE}="참쉬운 가족결합",CHOOSE({RANGE},CHOOSE(MIN({LINES},4)-1,2200,3300,4400),CHOOSE(MIN({LINES},4)-1,3300,5500,6600),CHOOSE(MIN({LINES},4)-1,4400,6600,8800)),0)',
    comment='→ 14_LGU_Chweyswun · 요금구간 × 회선수')
CHW_MOB_PER = r - 1

r = frow(ws, r, '  [참쉬운] 휴대폰 총 할인',
    f'=B{CHW_MOB_PER}*IF({BUNDLE}="참쉬운 가족결합",MIN({LINES},4),0)',
    comment='인당 × min(회선수,4)')
CHW_MOB_TOTAL = r - 1

r = frow(ws, r, '  [투게더] 인터넷 할인',
    f'=IF({BUNDLE}="투게더 결합",IF({SP}="100M",0,11000),0)',
    comment='→ 15_LGU_Together · 100M 결합불가 · 단독가 REPLACE')
TOG_INET = r - 1

r = frow(ws, r, '  [투게더] 휴대폰 인당 할인',
    f'=IF({BUNDLE}="투게더 결합",CHOOSE(MIN({LINES},4)-1,10000,14000,20000),0)',
    comment='→ 15_LGU_Together · 2/3/4+ 회선')
TOG_MOB_PER = r - 1

r = frow(ws, r, '  [투게더] 휴대폰 총 할인',
    f'=B{TOG_MOB_PER}*IF({BUNDLE}="투게더 결합",MIN({LINES},5),0)',
    comment='인당 × 회선수')
TOG_MOB_TOTAL = r - 1

r += 1
r = frow(ws, r, '  = 인터넷+TV 월 실납부 (결합 적용)',
    f'=(B{SOLO}-B{CHW_INET}-B{TOG_INET})+IF(B{HAS_TV}=0,0,B{TV_P}-B{TV_DC}+B{SETTOP})',
    LGU_FILL, True)
BUNDLE_FEE = r - 1

r = frow(ws, r, '  - 휴대폰 고지서 별도 차감',
    f'=-(B{CHW_MOB_TOTAL}+B{TOG_MOB_TOTAL})', LGU_FILL)

r = frow(ws, r, '🎉 혜택가 (최종 월요금)',
    f'=IF({BUNDLE}="결합없음",B{PATH_A},B{BUNDLE_FEE}-(B{CHW_MOB_TOTAL}+B{TOG_MOB_TOTAL}))',
    RESULT_FILL, True, 'e40981')
c = ws.cell(row=r-1, column=2)
c.font = Font(name='SF Mono', size=16, bold=True, color='e40981')

r += 1
c = ws.cell(row=r, column=1, value='🎁 부가')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666')
r += 1
r = frow(ws, r, '  설치비', f'=IF(B{HAS_TV}=0,36300,56100)', comment='→ 06_Install')
r = frow(ws, r, '  사은품', f'=IF(B{HAS_TV}=0,IF({SP}="100M",200000,230000),CHOOSE(MATCH({SP},{{"100M","500M","1G"}},0),400000,470000,470000))', comment='→ 07_Gift')

# ═══════════════════════════════════════════════
# 데이터 시트 (동일)
# ═══════════════════════════════════════════════
ws = wb.create_sheet('01_Internet')
for col in 'ABCDE': ws.column_dimensions[col].width = 16
r = title(ws, 1, '📡 인터넷 단독 요금')
r = table(ws, r, ['통신사', '프리픽스', '100M', '500M', '1G'], [
    ['SKT','SK',22000,33000,38500],['KT','KT',22000,33000,38500],['LG U+','LG',22000,33000,38500],
])

ws = wb.create_sheet('02_TV')
for col, w in zip('ABCDEFG', [10, 40, 10, 12, 12, 12, 15]):
    ws.column_dimensions[col].width = w
r = title(ws, 1, '📺 TV 상품 (3사)')
tv_all = []
for c_name, tvs in [('SKT', SKT_TVS), ('KT', KT_TVS), ('LG U+', LGU_TVS)]:
    for t in tvs[1:]:
        tv_all.append([c_name, t[0], t[1], t[2], t[3], t[2]-t[3]])
r = table(ws, r, ['통신사','TV 상품','채널수','단독가','결합할인','결합가'], tv_all)

ws = wb.create_sheet('03_TV_Internet')
for col in 'ABCDE': ws.column_dimensions[col].width = 16
r = title(ws, 1, '🔗 TV 결합 시 인터넷 요금')
r = table(ws, r, ['통신사','WiFi','100M','500M','1G'], [
    ['SKT','미포함',19800,27500,33000],['SKT','포함',22000,28600,34100],
    ['KT','미포함',22000,27500,33000],['KT','포함 (1G=동일)',23100,28600,33000],
    ['LG U+','공통',22000,27500,33000],
])

ws = wb.create_sheet('04_SetTop')
for col, w in zip('ABCDEFG', [10,20,28,12,10,10,50]): ws.column_dimensions[col].width = w
r = title(ws, 1, '📦 셋톱박스')
c = ws.cell(row=r, column=1, value='💬 기본값(isDefault): 매장 판매 1위 + 가격/기능 균형 · 가입 시 자동 설정 · 요금 계산 기본')
c.font = Font(name='Noto Sans KR', size=10, italic=True, color='92400e')
c.fill = PatternFill('solid', fgColor='fef3c7')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); r += 2
r = table(ws, r, ['통신사','ID','모델명','월 임대료','기본','활성','💬 설명'], [
    ['SKT','smart3','스마트3',4400,'O','O','✅ 기본 · SKT 표준 · OTT 대부분 지원 · 매장 판매 1위'],
    ['SKT','smart3-mini','스마트3 미니',4400,'X','O','저가형 · 넷플릭스 미지원'],
    ['SKT','ai-nugu','AI NUGU',6600,'X','O','AI 음성인식 특화'],
    ['SKT','sound-max','사운드 맥스',8800,'X','O','사운드바 일체 · OTT 거의 불가'],
    ['SKT','apple-tv','애플TV',6600,'X','O','아이폰 연동'],
    ['KT','genie-3','기가지니3',4400,'O','O','✅ 기본 · KT 대표 · AI 블루투스 일체 · 선호 1위'],
    ['KT','genie-a','기가지니A',3300,'X','O','저가형 (3,300원)'],
    ['KT','soundbar','지니TV 올인원 사운드바',8800,'X','O','사운드바+공유기 일체'],
    ['LG U+','uhd4','U+tv UHD4',4400,'O','O','✅ 기본 · AI 음향기술'],
    ['LG U+','soundbar-black','U+tv 사운드바 블랙',6600,'X','O','돌비비전 지원'],
])

ws = wb.create_sheet('05_WiFi')
for col, w in zip('ABCDEFGHI', [10,20,28,10,10,10,8,8,50]): ws.column_dimensions[col].width = w
r = title(ws, 1, '📡 WiFi')
c = ws.cell(row=r, column=1, value='💬 기본값: 속도별 단가가 낮은 표준 모델 · 가입 시 자동 지급 · 요금 계산 대상')
c.font = Font(name='Noto Sans KR', size=10, italic=True, color='92400e')
c.fill = PatternFill('solid', fgColor='fef3c7')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); r += 2
r = table(ws, r, ['통신사','ID','모델명','100M','500M','1G','기본','활성','💬 설명'], [
    ['SKT','giga-wifi','GIGA WiFi',1100,1100,1100,'O','O','✅ 기본 · SKT 표준 · 속도 균일 1,100원'],
    ['SKT','giga-wifi-6','GIGA WiFi 6',1100,1100,1100,'X','O','WiFi-6 · 같은 요금'],
    ['SKT','giga-wifi-prem','GIGA WiFi 프리미엄',5500,5500,5500,'X','O','최대 1.7Gbps · 고급형'],
    ['SKT','wings','윙즈',1650,1650,1650,'X','O','WiFi 증폭기'],
    ['KT','wave2','KT GIGA WAVE2',1100,1100,0,'O','O','✅ 기본 · KT 표준 · 1G 무료'],
    ['KT','home-ax','GIGA WIFI 홈AX',0,1100,0,'X','O','WiFi-6 · 100M/1G 무료'],
    ['KT','buddy','GIGA WIFI BUDDY',1650,1650,1650,'X','O','증폭기 (홈AX 호환)'],
    ['KT','prem-24','GIGA WIFI 프리미엄 2.4',4400,4400,4400,'X','O','고급형 10Gbps'],
    ['KT','prem-24-6e','프리미엄 2.4 (6E)',4400,4400,4400,'X','O','6E 지원'],
    ['KT','prem-48','프리미엄 4.8',4400,4400,4400,'X','O','4.8GHz 대역'],
    ['LG U+','giga-wifi','기가와이파이',0,0,0,'O','O','✅ 기본 · 전속도 무료'],
    ['LG U+','giga-wifi-6','기가와이파이6',0,0,0,'X','O','WiFi-6 · 동일 무료'],
    ['LG U+','giga-wifi-mesh','기가와이파이 메쉬',0,0,0,'X','O','메쉬 · 500M+ 선택'],
])

ws = wb.create_sheet('06_Install')
for col in 'ABCD': ws.column_dimensions[col].width = 18
r = title(ws, 1, '🔧 설치비')
r = table(ws, r, ['통신사','단독','결합','비고'], [['SKT',36300,56100,'주말 +25%'],['KT',36000,56200,'주말 +25%'],['LG U+',36300,56100,'주말 +25%']])

ws = wb.create_sheet('07_Gift')
for col in 'ABCDE': ws.column_dimensions[col].width = 14
r = title(ws, 1, '🎁 사은품')
r = table(ws, r, ['통신사','구분','100M','500M','1G'], [
    ['SKT','solo',110000,170000,170000],['SKT','combo',400000,430000,490000],
    ['KT','solo',90000,140000,140000],['KT','combo',370000,450000,450000],
    ['LG U+','solo',200000,230000,230000],['LG U+','combo',400000,470000,470000],
])

ws = wb.create_sheet('08_Cards')
for col, w in zip('ABCDE', [10,14,40,14,30]): ws.column_dimensions[col].width = w
r = title(ws, 1, '💳 제휴카드 (26종)')
cards = [
    ['SKT','롯데카드','SK브로드밴드 B롯데카드','50만','-10,000원'],
    ['SKT','삼성카드','SK브로드밴드 삼성카드','30만','-7,000원'],
    ['KT','KB국민','KT DC Plus 국민카드','30만','-7,000원'],
    ['KT','현대','KT-현대카드M Edition3','30만','-13,000원'],
    ['KT','현대','KT-현대카드M Edition3 (2.0)','100만','-22,000원'],
    ['KT','신한','KT 신한 체크카드','30만','3,000원 캐시백'],
    ['KT','신한','KT 가족만족 DC 신한카드','30만','-7,000원'],
    ['KT','신한','KT 으랏차차 신한카드','50만','-12,000원'],
    ['KT','IBK','olleh super DC IBK카드','30만','-7,000원'],
    ['KT','IBK','KT 으랏차차 IBK카드','자동납부','5% 청구할인'],
    ['KT','삼성','KT 삼성카드','30만','-7,000원'],
    ['KT','우리','KT Plus 우리카드','40만','-10,000원'],
    ['KT','우리','KT 36 Plus 우리카드','40만','-8,000원'],
    ['KT','하나','KT DC Plus 더 심플 하나카드','30만','-10,000원'],
    ['KT','NH농협','KT 할부 Plus NH농협카드','40만','-5,000원'],
    ['KT','롯데','KT DC Plus 롯데카드','40만','-10,000원'],
    ['KT','비씨','KT SUPER DC BC 바로카드','40만','-5,000원'],
    ['KT','비씨','KT DC Plus BC 바로카드','30만','-7,000원'],
    ['KT','케이뱅크','KT멤버십x케이뱅크','20만','5% 캐시백'],
    ['LG U+','삼성카드','LG U+ 삼성카드','30만','-7,000원'],
    ['LG U+','현대카드','LG U+ 현대카드M Edition3','50만','-15,000원'],
    ['LG U+','하나카드','더 심플 하나카드','30만','-10,000원'],
    ['LG U+','하나카드','LG U+Family 하나카드','30만','통신료 25%'],
    ['LG U+','신한카드','LG U+ 사장님 통할인','70만','-10,000원'],
    ['LG U+','롯데카드','LG U+ x LOCA','30만','-10,000원'],
    ['LG U+','NH카드','NH올원 LG U+ 카드','30만','-9,000원'],
]
r = table(ws, r, ['통신사','카드사','카드명','실적','할인'], cards)

# 결합할인 시트들
ws = wb.create_sheet('10_SKT_Bundle')
for col in 'ABCDEF': ws.column_dimensions[col].width = 18
r = title(ws, 1, '🏷️ SKT 요즘가족결합 (완전 명세)', 'ef4444')

# ① 자격 조건
c = ws.cell(row=r, column=1, value='① 자격 조건')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='ef4444'); r += 1
for line in [
    '▶ 같은 명의자의 인터넷+휴대폰 1대 이상 필수',
    '▶ 가족/동거인/친구 등 지인 결합 허용 (최대 5회선)',
    '▶ 적용 방식: 단독가 REPLACE (기존 요즘우리집 인터넷 할인 대체)',
    '▶ TV 결합 시: IPTV 추가 -1,100원 중복',
    '▶ 인터넷 최대 2회선 + 휴대폰 최대 5회선',
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name='Noto Sans KR', size=10)
    c.fill = PatternFill('solid', fgColor='fef3c7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
r += 1

# ② 할인 표
c = ws.cell(row=r, column=1, value='② 할인 테이블 (속도 × 회선수)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='ef4444'); r += 1
skt_rows = []
mob_100 = {1:3500,2:7000,3:18000,4:18000,5:18000}
mob_500 = {1:3500,2:7000,3:18000,4:24000,5:24000}
inet_map = {'100M':4400,'500M':11000,'1G':13200}
for sp in ['100M','500M','1G']:
    for ln in [1,2,3,4,5]:
        mob = mob_100[ln] if sp=='100M' else mob_500[ln]
        inet = inet_map[sp]
        skt_rows.append([sp, ln, mob, inet, 1100, mob+inet+1100])
r = table(ws, r, ['속도','회선수','휴대폰 (총합)','인터넷','IPTV 추가','총할인'], skt_rows, SKT_FILL)
r += 1

# 인당 할인 참고
c = ws.cell(row=r, column=1, value='📌 휴대폰 인당 할인 (개발 참조)')
c.font = Font(name='Noto Sans KR', size=11, bold=True, color='666666'); r += 1
r = table(ws, r, ['속도','1회선','2회선','3회선','4회선','5회선'], [
    ['100M (인당)',3500,3500,6000,4500,3600],
    ['500M·1G (인당)',3500,3500,6000,6000,4800],
])
r += 1

# ③ 계산 알고리즘
c = ws.cell(row=r, column=1, value='③ 계산 알고리즘 (의사코드)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='ef4444'); r += 1
for line in [
    '// 입력: speed, tvIdx, wifi(bool), lines(1~5)',
    '',
    '1. 단독가 기준으로 재계산 (요즘우리집 REPLACE):',
    '   netSingle = internet[speed] + (wifi ? wifiCost : 0)  // SKT wifiCost=1100',
    '   famInet = bundle.family.internet[speed]',
    '   inetAfter = netSingle - famInet',
    '',
    '2. TV 결합 시:',
    '   famIptv = hasTv ? bundle.family.iptv : 0  // iptv=1100',
    '   tvAfter = tv[tvIdx].p - tv[tvIdx].dc - famIptv',
    '   monthlyFee = inetAfter + tvAfter + setTop',
    '',
    '3. 휴대폰 별도 차감 (월):',
    '   perPerson = bundle.family.mobilePerBySpeed[speed][lines]',
    '   mobileDiscount = perPerson × lines  // 고지서에서 차감',
    '',
    '4. 혜택가:',
    '   finalFee = monthlyFee - mobileDiscount',
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name='SF Mono', size=10, color='1a2744')
    c.fill = PatternFill('solid', fgColor='f5f5f7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
r += 1

# ④ 검증 예시
c = ws.cell(row=r, column=1, value='④ 검증 예시 — 500M + B tv 올 + WiFi X + 3회선')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='ef4444'); r += 1
r = table(ws, r, ['단계','계산','값'], [
    ['단독가','33,000 + 0','33,000'],
    ['요즘가족 인터넷 할인','-11,000','-11,000'],
    ['인터넷 실질','33,000 - 11,000','22,000'],
    ['TV 기본 (B tv 올)','18,700','18,700'],
    ['TV 결합할인 + IPTV','-2,200 + -1,100','-3,300'],
    ['TV 실질','18,700 - 3,300','15,400'],
    ['셋톱 (스마트3)','4,400','4,400'],
    ['= 월 실납부','22,000 + 15,400 + 4,400','41,800'],
    ['휴대폰 3회선 × 6,000','-18,000','-18,000'],
    ['🎉 혜택가','41,800 - 18,000','23,800'],
], SKT_FILL)
r += 1

# ⑤ 참고자료
c = ws.cell(row=r, column=1, value='⑤ 참고자료')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='ef4444'); r += 1
rc = ws.cell(row=r, column=1, value='🔗 백메가 SK 상품 전체 안내 (결합할인 원본)')
rc.hyperlink = 'https://www.100mb.kr/01_product/sk.php'
rc.font = Font(name='Noto Sans KR', size=11, color='2563eb', underline='single', bold=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
rc2 = ws.cell(row=r, column=1, value='    URL: https://www.100mb.kr/01_product/sk.php')
rc2.font = Font(name='SF Mono', size=9, color='666666')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

ws = wb.create_sheet('11_KT_Total')
for col in 'ABCDE': ws.column_dimensions[col].width = 17
r = title(ws, 1, '🏷️ KT 총액결합 (STACK)', '2563eb')
ranges_t = ['22K 이하','22K 이상','64.9K 이상','108.9K 이상','141.9K 이상','174.9K 이상']
inet_t = {'100M':[1650,3300,5500,5500,5500,5500],'500M':[2200,5500,5500,5500,5500,5500],'1G':[2200,5500,5500,5500,5500,5500]}
mob_t = {'100M':[0,0,3300,14300,18700,23100],'500M':[0,0,5500,16610,22110,27610],'1G':[0,0,5500,16610,22110,27610]}
rows = []
for sp in ['100M','500M','1G']:
    for i in range(6):
        rows.append([sp, ranges_t[i], inet_t[sp][i], mob_t[sp][i], inet_t[sp][i]+mob_t[sp][i]])
r = table(ws, r, ['속도','구간','인터넷','휴대폰','합계'], rows, KT_FILL)

ws = wb.create_sheet('12_KT_Fixed')
for col in 'ABCD': ws.column_dimensions[col].width = 18
r = title(ws, 1, '🏷️ KT 정액결합', '2563eb')
r = table(ws, r, ['구간','인터넷','휴대폰','합계'], [
    ['37K 이하',5500,0,5500],['37K 이상',5500,3000,8500],
    ['61K 이상',5500,5000,10500],['77K 이상',5500,7000,12500]
], KT_FILL)

ws = wb.create_sheet('13_KT_Premium')
for col, w in zip('ABCDEF', [14,44,14,14,14,30]): ws.column_dimensions[col].width = w
r = title(ws, 1, '💎 KT 프리미엄 가족결합 (완전 명세서)', 'd97706')

# ─── 1. 자격 조건 ───
c = ws.cell(row=r, column=1, value='① 자격 조건')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='d97706'); r += 1
for line in [
    '▶ 필수: 77,000원↑ 요금제 휴대폰 2회선 이상 결합',
    '▶ 대표명의자 1명 + 구성원 최대 N명 (총 5회선까지)',
    '▶ 자격 미충족 시: 총액결합만 가능 (프리미엄 선택 불가)',
    '▶ 77,000원↑ 요금제가 1회선뿐 (대표자 본인) → 프리미엄 불가',
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name='Noto Sans KR', size=10)
    c.fill = PatternFill('solid', fgColor='fef3c7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
r += 1

# ─── 2. 분배 규칙 ───
c = ws.cell(row=r, column=1, value='② 회선별 분배 규칙 (누가 어디 속하는가)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='d97706'); r += 1
r = table(ws, r, ['회선 종류','조건','귀속 할인','비고'], [
    ['대표명의자','무조건 (총대)','총액결합할인','프리미엄 불가 · 인터넷 -5,500 대체 가능'],
    ['구성원 (77K↑)','선택 가능','프리미엄 가족결합','요금제별 고정 dc × 회선수'],
    ['구성원 (77K↑)','선택 안 함','총액결합할인','합산에 포함 (구간 상승 효과)'],
    ['구성원 (77K미만)','무조건','총액결합할인','합산에 포함'],
])
r += 1

# ─── 3. 요금제 카탈로그 ───
c = ws.cell(row=r, column=1, value='③ 요금제 카탈로그 (원본 planCatalog 14종)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='d97706'); r += 1
r = table(ws, r, ['월정액 (v)','요금제명 (label)','prem 자격','할인액 (dc)','비율'], [
    [0,'— 회선 없음 —','X',0,'-'],
    [32890,'32,890원 (LTE 데이터선택)','X',0,'-'],
    [47000,'47,000원','X',0,'-'],
    [55000,'55,000원','X',0,'-'],
    [65000,'65,000원','X',0,'-'],
    [77000,'77,000원 (임계값)','✓',19250,'≈25%'],
    [80000,'80,000원 (5G 슈퍼플랜 베이직)','✓',20000,'25%'],
    [87890,'87,890원 (데이터선택 87.8)','✓',22000,'≈25%'],
    [89000,'89,000원 (데이터ON 프리미엄)','✓',22250,'25%'],
    [90000,'90,000원 (슈퍼플랜 베이직 Plus/초이스)','✓',22500,'25%'],
    [100000,'100,000원 (슈퍼플랜 스페셜)','✓',25000,'25%'],
    [109000,'109,000원 (데이터선택 109)','✓',27500,'≈25%'],
    [110000,'110,000원 (슈퍼플랜 스페셜 Plus/초이스)','✓',27500,'25%'],
    [130000,'130,000원 (슈퍼플랜 프리미엄/초이스)','✓',32500,'25%'],
], PatternFill('solid', fgColor='fef3c7'))
r += 1

# ─── 4. 계산 알고리즘 (의사코드) ───
c = ws.cell(row=r, column=1, value='④ 계산 알고리즘 (의사코드)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='d97706'); r += 1
algo_lines = [
    '// 입력: members: Array<{ isRep: boolean, planV: number, usePrem?: boolean }>',
    '//      speed: "100M" | "500M" | "1G"',
    '',
    '1. 자격 체크:',
    '   highPlanCount = members.filter(m => m.planV >= 77000).length',
    '   IF highPlanCount < 2 THEN 프리미엄 불가 (총액결합만 사용)',
    '',
    '2. 회선 분배:',
    '   premMembers = members.filter(m => !m.isRep && m.usePrem && m.planV >= 77000)',
    '   totalMembers = members.filter(m => m.isRep || !m.usePrem || m.planV < 77000)',
    '   IF premMembers.length == 0 THEN 프리미엄 불가 (총액결합만 사용)',
    '',
    '3. 총액결합 합산액:',
    '   totalSum = sum(totalMembers.map(m => m.planV))',
    '',
    '4. 총액결합 구간 판정:',
    '   rngIdx = 0 ~ 5',
    '   if totalSum >= 174900: rngIdx = 5',
    '   else if totalSum >= 141900: rngIdx = 4',
    '   else if totalSum >= 108900: rngIdx = 3',
    '   else if totalSum >= 64900:  rngIdx = 2',
    '   else if totalSum >= 22000:  rngIdx = 1',
    '   else: rngIdx = 0',
    '',
    '5. 할인 계산:',
    '   inetFixedDc = 5500  // 인터넷 또는 대표자 휴대폰 (기본: 대표자 휴대폰)',
    '   totalMobDc = kt.total.mobile[speed][rngIdx]',
    '   premTotalDc = sum(premMembers.map(m => planCatalog[m.planV].dc))',
    '',
    '6. 최종:',
    '   totalDiscount = inetFixedDc + totalMobDc + premTotalDc',
]
for line in algo_lines:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name='SF Mono', size=10, color='1a2744')
    c.fill = PatternFill('solid', fgColor='f5f5f7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
r += 1

# ─── 5. 검증 예시 1 ───
c = ws.cell(row=r, column=1, value='⑤ 검증 예시 1 — 100M + 3회선 (모두 89,000원)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='d97706'); r += 1
r = table(ws, r, ['단계','구성','계산','할인액'], [
    ['회선 1','대표자 89K (총액)','자동','대표자는 총액결합 대상'],
    ['회선 2','구성원 89K (프리미엄)','체크','프리미엄 대상'],
    ['회선 3','구성원 89K (프리미엄)','체크','프리미엄 대상'],
    ['총액 합산','대표자 1명 89,000원','rngIdx=2 (64.9K↑)','-'],
    ['인터넷','대표자 휴대폰에 -5,500','고정','-5,500'],
    ['대표자 총액결합','100M mobile[2]','3,300','-3,300'],
    ['구성원 프리미엄','89K × 2명','22,250 × 2','-44,500'],
    ['🎉 총 할인','5,500 + 3,300 + 44,500','','-53,300'],
    ['비교 (총액결합만)','3×89K=267K → idx 5','mobile[5]=23,100 + 인터넷 5,500','-28,600'],
    ['차액','프리미엄 선택 시 추가 혜택','','+24,700'],
], PatternFill('solid', fgColor='fef3c7'))
r += 1

# ─── 6. 검증 예시 2 ───
c = ws.cell(row=r, column=1, value='⑥ 검증 예시 2 — 100M + 4회선 (2×89K + 2×32.89K)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='d97706'); r += 1
r = table(ws, r, ['단계','구성','계산','할인액'], [
    ['회선 1','대표자 89K','총액','-'],
    ['회선 2','구성원 89K (프리미엄)','체크','프리미엄 대상'],
    ['회선 3','구성원 32.89K','77K미만 → 자동 총액','-'],
    ['회선 4','구성원 32.89K','77K미만 → 자동 총액','-'],
    ['총액 합산','89 + 32.89×2 = 154,780','rngIdx=4 (141.9K↑)','-'],
    ['인터넷','고정','','-5,500'],
    ['대표자+일반 총액결합','100M mobile[4]','18,700','-18,700'],
    ['프리미엄 89K','1명','22,250','-22,250'],
    ['🎉 총 할인','5,500 + 18,700 + 22,250','','-46,450'],
], PatternFill('solid', fgColor='fef3c7'))
r += 1

# ─── 7. 엣지 케이스 ───
c = ws.cell(row=r, column=1, value='⑦ 엣지 케이스 처리')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='d97706'); r += 1
r = table(ws, r, ['케이스','처리'], [
    ['77K+ 요금제가 1명만 (대표자만)','프리미엄 선택 불가 → 총액결합으로만 계산'],
    ['77K+ 구성원 중 프리미엄 체크 0명','프리미엄 미적용 (전원 총액)'],
    ['대표자 요금제가 77K 미만','가능 · 대표자는 항상 총액결합'],
    ['모든 구성원이 77K 미만','프리미엄 불가 (자격 미달)'],
    ['1회선만 결합','프리미엄 불가 (2회선 이상 필수)'],
    ['프리미엄 구성원 5명+','현실적으로 5회선까지 결합 가능 (KT 규정)'],
])
r += 1

# ─── 8. 참고자료 ───
c = ws.cell(row=r, column=1, value='⑧ 참고자료 (원본 출처)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='d97706'); r += 1
ref_cell = ws.cell(row=r, column=1, value='🔗 백메가 블로그 — KT 프리미엄 가족결합 상세 해설 (원본 소스)')
ref_cell.hyperlink = 'https://www.100mb.kr/bbs/board.php?bo_table=information&wr_id=11986'
ref_cell.font = Font(name='Noto Sans KR', size=11, color='2563eb', underline='single', bold=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
ref_cell2 = ws.cell(row=r, column=1, value='    URL: https://www.100mb.kr/bbs/board.php?bo_table=information&wr_id=11986')
ref_cell2.font = Font(name='SF Mono', size=9, color='666666')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
ref_note = ws.cell(row=r, column=1, value='    · 자격 조건, 대표자/구성원 분배 규칙, 실제 계산 예시 2건 원본')
ref_note.font = Font(name='Noto Sans KR', size=10, color='666666', italic=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

ws = wb.create_sheet('14_LGU_Chweyswun')
for col in 'ABCDE': ws.column_dimensions[col].width = 18
r = title(ws, 1, '🏷️ LG U+ 참쉬운가족결합 (완전 명세)', 'e40981')

c = ws.cell(row=r, column=1, value='① 자격 조건')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
for line in [
    '▶ 휴대폰 2회선부터 결합 (1회선 불가)',
    '▶ 알뜰폰 결합 가능 (추가 할인 대당 -440원)',
    '▶ 휴대폰 최대 10대 + 인터넷 최대 3대',
    '▶ 적용 방식: 단독가 REPLACE',
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name='Noto Sans KR', size=10)
    c.fill = PatternFill('solid', fgColor='fef3c7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
r += 1

c = ws.cell(row=r, column=1, value='② 인터넷 할인 (3년 약정 기준)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
r = table(ws, r, ['구분','100M','500M','1G'], [['인터넷 할인',5500,9900,13200]], LGU_FILL)
r += 1

c = ws.cell(row=r, column=1, value='③ 휴대폰 인당 할인 매트릭스 (요금구간 × 회선수)')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
r = table(ws, r, ['요금구간','2회선','3회선','4+회선'], [
    ['69K 미만',2200,3300,4400],['69K 이상',3300,5500,6600],['88K 이상',4400,6600,8800],
], LGU_FILL)
r += 1

c = ws.cell(row=r, column=1, value='④ 계산 알고리즘')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
for line in [
    '// 입력: speed, tvIdx, planRange(1~3), lines(2~4+)',
    '1. lguInet = bundle.chweyswun.internet[speed]',
    '2. netSingle = internet[speed]  // LGU+ WiFi 전속도 무료',
    '3. inetAfter = netSingle - lguInet  // REPLACE 방식',
    '4. 휴대폰: idx = min(lines, 4) - 2  // 2회선=0 / 3회선=1 / 4+=2',
    '   perPerson = bundle.chweyswun.mobile[planRange-1][idx]',
    '   mobileDc = perPerson × min(lines, 4)',
    '5. TV: tvAfter = tv[tvIdx].p - tv[tvIdx].dc',
    '6. monthlyFee = inetAfter + tvAfter + setTop(4,400)',
    '7. finalFee = monthlyFee - mobileDc',
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name='SF Mono', size=10, color='1a2744')
    c.fill = PatternFill('solid', fgColor='f5f5f7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
r += 1

c = ws.cell(row=r, column=1, value='⑤ 참고자료')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
rc = ws.cell(row=r, column=1, value='🔗 백메가 LG U+ 상품 전체 안내 (결합할인 원본 소스)')
rc.hyperlink = 'https://www.100mb.kr/01_product/lg.php'
rc.font = Font(name='Noto Sans KR', size=11, color='2563eb', underline='single', bold=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
rc2 = ws.cell(row=r, column=1, value='    URL: https://www.100mb.kr/01_product/lg.php')
rc2.font = Font(name='SF Mono', size=9, color='666666')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

ws = wb.create_sheet('15_LGU_Together')
for col in 'ABCD': ws.column_dimensions[col].width = 20
r = title(ws, 1, '🏷️ LG U+ 투게더 결합 (85,000원↑ 고가요금제)', 'e40981')

c = ws.cell(row=r, column=1, value='① 자격 조건')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
for line in [
    '▶ 5G 85,000원 이상 요금제 필수',
    '▶ 500M 이상 인터넷 (100M 결합 불가)',
    '▶ 휴대폰 최대 5회선 + 인터넷 최대 5회선',
    '▶ 선택약정 25% 중복 가능',
    '▶ 적용 방식: 단독가 REPLACE',
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name='Noto Sans KR', size=10)
    c.fill = PatternFill('solid', fgColor='fef3c7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
r += 1

c = ws.cell(row=r, column=1, value='② 인터넷 할인')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
r = table(ws, r, ['구분','100M','500M','1G'], [['인터넷','❌ 불가',11000,11000]], LGU_FILL)
r += 1

c = ws.cell(row=r, column=1, value='③ 휴대폰 인당 할인')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
r = table(ws, r, ['회선수','인당 할인'], [['2회선',10000],['3회선',14000],['4~5회선',20000]], LGU_FILL)
r += 1
c = ws.cell(row=r, column=1, value='※ 추가 혜택: 청소년 가족할인 -10,000 / 시그니처 가족할인 자녀 1대 최대 -33,000')
c.font = Font(name='Noto Sans KR', size=9, italic=True, color='666666')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2

c = ws.cell(row=r, column=1, value='④ 참고자료')
c.font = Font(name='Noto Sans KR', size=13, bold=True, color='e40981'); r += 1
rc = ws.cell(row=r, column=1, value='🔗 백메가 LG U+ 상품 안내')
rc.hyperlink = 'https://www.100mb.kr/01_product/lg.php'
rc.font = Font(name='Noto Sans KR', size=11, color='2563eb', underline='single', bold=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

# Schema & Formula
ws = wb.create_sheet('20_Schema')
ws.column_dimensions['A'].width = 90
r = title(ws, 1, '📐 TypeScript 인터페이스')
for line in [
    'interface CarrierData {',
    '  name, prefix;',
    '  internet: Record<"100M"|"500M"|"1G", number>;',
    '  wifiCost? | wifiPrice?: Record<Speed, number>;',
    '  setTopOptions: SetTop[]; wifiOptions: WiFi[];',
    '  tvInternetNoWifi, tvInternetWithWifi: Record<Speed, number>;',
    '  tv: TVProduct[];',
    '  install: { solo, combo };',
    '  gift: { solo: SpeedMap, combo: SpeedMap };',
    '  bundle: BundleConfig;  // 3사별 상이',
    '}',
    '',
    '// SKT: bundle.family { internet, iptv, mobilePerBySpeed }',
    '// KT:  bundle.{ranges_total, total, ranges_fixed, fixed, premium}',
    '// LGU: bundle.{chweyswun, together}',
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name='SF Mono', size=10, color='1a2744'); r += 1

ws = wb.create_sheet('30_Formula')
ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 70
r = title(ws, 1, '🧮 계산 공식')
c = ws.cell(row=r, column=1, value='⚠️ 3사 결합할인 적용 방식 차이')
c.font = Font(name='Noto Sans KR', size=12, bold=True, color='dc2626')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 1
for line in [
    'SKT 요즘가족결합: 단독가 REPLACE (요즘우리집 대체) → tvInternet 사용 X',
    'KT 총액/정액: 기본 TV결합 + 추가할인 STACK (중복)',
    'KT 프리미엄: 대표자+77K미만은 총액결합 / 77K↑ 구성원은 프리미엄 (독립)',
    'LG U+ 참쉬운: 단독가 REPLACE',
    'LG U+ 투게더: 단독가 REPLACE + 100M 결합 불가',
]:
    c = ws.cell(row=r, column=1, value='• ' + line)
    c.font = Font(name='Noto Sans KR', size=10)
    c.fill = PatternFill('solid', fgColor='fef3c7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 1

# 저장
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f'✅ {OUT} ({len(wb.sheetnames)} 시트)')
print('   시트:', ', '.join(wb.sheetnames))
os.makedirs(os.path.dirname(SCRIPT_OUT), exist_ok=True)
shutil.copy(__file__, SCRIPT_OUT)
print(f'✅ 스크립트 보관: {SCRIPT_OUT}')
